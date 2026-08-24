#!/usr/bin/env python3

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_DOI = "10.5281/zenodo.10678852"
SOURCE_URL = "https://zenodo.org/records/11498333"
WORKBOOKS = {
    "assertion_failure": "Test_Failure_Data.xlsx",
    "exception_failure": "Test_Errors_Data.xlsx",
}
GROUP_FIELDS = (
    "Dependency",
    "Current Version",
    "Breaking Version",
    "Project Name",
    "Directory",
)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def load_rows(path: Path, failure_kind: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    values = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(values)]
    rows: list[dict[str, Any]] = []
    for row_number, values_row in enumerate(values, start=2):
        row = dict(zip(headers, values_row, strict=True))
        row["_failure_kind"] = failure_kind
        row["_source_row"] = row_number
        row["_source_workbook"] = path.name
        rows.append(row)
    workbook.close()
    return rows


def unique_text(rows: list[dict[str, Any]], field: str) -> list[str]:
    return sorted({text(row.get(field)) for row in rows if text(row.get(field))})


def excerpt(value: Any, limit: int = 500) -> str:
    normalized = text(value).replace("\r\n", "\n")
    return normalized if len(normalized) <= limit else normalized[:limit] + "..."


def build_candidate(index: int, rows: list[dict[str, Any]]) -> dict[str, Any]:
    first = rows[0]
    assertion_rows = [row for row in rows if row["_failure_kind"] == "assertion_failure"]
    exception_rows = [row for row in rows if row["_failure_kind"] == "exception_failure"]
    source_locations = [
        {
            "workbook": row["_source_workbook"],
            "sheet": "Sheet1",
            "row": row["_source_row"],
            "record_id": row.get("Record ID"),
            "current_dependency_record_id": row.get("current_dep_notransitive_id"),
        }
        for row in sorted(
            rows,
            key=lambda item: (item["_source_workbook"], item["_source_row"]),
        )
    ]
    representative = next((row for row in rows if text(row.get("Error Log"))), first)
    return {
        "candidate_id": f"fse2024-behavioral-{index:04d}",
        "candidate_status": "executed_breakage_lead_only",
        "dependency": {
            "coordinate": text(first.get("Dependency")),
            "current_version": text(first.get("Current Version")),
            "previous_version": unique_text(rows, "Previous Version"),
            "breaking_version": text(first.get("Breaking Version")),
        },
        "client": {
            "maven_artifact": text(first.get("Project Name")),
            "repository_directory_hint": text(first.get("Directory")),
            "build_java_versions": unique_text(rows, "Build Java Version"),
            "dependency_scopes": unique_text(rows, "Scope"),
            "repository": None,
            "revision": None,
        },
        "observed_failure": {
            "assertion_records": len(assertion_rows),
            "exception_records": len(exception_rows),
            "test_methods": unique_text(rows, "Test Method"),
            "error_types": unique_text(rows, "Error Type"),
            "assertion_origins": unique_text(assertion_rows, "Assertion Origin"),
            "exception_origins": unique_text(exception_rows, "Exception Origin"),
            "exception_types": unique_text(exception_rows, "Exception Type"),
            "exception_categories": unique_text(exception_rows, "Exception Category"),
            "representative_error_excerpt": excerpt(representative.get("Error Log")),
        },
        "source_evidence": {
            "doi": SOURCE_DOI,
            "record_url": SOURCE_URL,
            "locations": source_locations,
        },
        "missing_for_causal_admission": [
            "Recover the exact source repository and revisions for the two dependency versions.",
            "Isolate the behavioral change to one source commit or pull request rather than the full release delta.",
            "Recover the exact client repository revision used by the published execution.",
            "Find and isolate a maintainer-authored client repair for the same failure contract.",
            "Replay A0, A1, and A2 before treating the relation as a positive label.",
            "Find an independently justified compatible source change and covered distractors for A3.",
        ],
        "label_boundary": (
            "The artifact records a client test regression after a dependency update. It does not "
            "identify a maintainer repair or establish a complete cross-repository impact set."
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--metadata", type=Path, required=True)
    parser.add_argument("--observation-date", required=True)
    args = parser.parse_args()

    all_rows: list[dict[str, Any]] = []
    source_counts: dict[str, int] = {}
    for failure_kind, workbook_name in WORKBOOKS.items():
        rows = load_rows(args.input_dir / workbook_name, failure_kind)
        all_rows.extend(rows)
        source_counts[workbook_name] = len(rows)

    groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        key = tuple(text(row.get(field)) for field in GROUP_FIELDS)
        groups[key].append(row)

    ordered_groups = sorted(groups.values(), key=lambda rows: tuple(text(rows[0].get(field)) for field in GROUP_FIELDS))
    candidates = [build_candidate(index, rows) for index, rows in enumerate(ordered_groups, start=1)]

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as handle:
        for candidate in candidates:
            handle.write(json.dumps(candidate, ensure_ascii=False, sort_keys=True) + "\n")

    dependency_coordinates = {text(row.get("Dependency")) for row in all_rows}
    client_artifacts = {text(row.get("Project Name")) for row in all_rows}
    client_directories = {text(row.get("Directory")) for row in all_rows}
    dependency_client_pairs = {
        (text(row.get("Dependency")), text(row.get("Project Name")), text(row.get("Directory")))
        for row in all_rows
    }
    metadata = {
        "observation_date": args.observation_date,
        "source": {
            "doi": SOURCE_DOI,
            "record_url": SOURCE_URL,
            "workbook_rows": source_counts,
        },
        "grouping_fields": list(GROUP_FIELDS),
        "source_failure_rows": len(all_rows),
        "candidate_dependency_updates": len(candidates),
        "unique_dependency_coordinates": len(dependency_coordinates),
        "unique_client_artifacts": len(client_artifacts),
        "unique_client_directory_hints": len(client_directories),
        "unique_dependency_client_pairs": len(dependency_client_pairs),
        "candidates_with_assertion_failures": sum(
            candidate["observed_failure"]["assertion_records"] > 0 for candidate in candidates
        ),
        "candidates_with_exception_failures": sum(
            candidate["observed_failure"]["exception_records"] > 0 for candidate in candidates
        ),
        "candidates_with_both_failure_kinds": sum(
            candidate["observed_failure"]["assertion_records"] > 0
            and candidate["observed_failure"]["exception_records"] > 0
            for candidate in candidates
        ),
        "limitations": [
            "The workbooks record executed dependency-update regressions but omit exact repository URLs and revisions.",
            "A breaking release can contain multiple source changes and is not yet a single-change Marshal input.",
            "A test failure does not by itself identify the repository that should be changed.",
            "The artifact does not provide maintainer repair commits or a complete unaffected-client set.",
            "Candidates must not be scored until repository recovery and A0/A1/A2 replay succeed.",
        ],
    }
    args.metadata.parent.mkdir(parents=True, exist_ok=True)
    args.metadata.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
